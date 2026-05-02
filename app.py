import streamlit as st
import base64
import json
import io
import os
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Electricity Bill Extractor", page_icon="⚡")
st.title("⚡ Electricity Bill Extractor")
st.caption("Upload a bill image → AI extracts all data → Download as Excel")

SYSTEM_PROMPT = """You are an expert at reading electricity bills.
Extract data from the bill image and return ONLY a valid JSON object with these exact keys:

{
  "consumer_name": "",
  "consumer_no": "",
  "fixed_charges": "",
  "sanct_load_kw": "",
  "connection_type": "",
  "contract_demand_kva": "",
  "solar_panel_used": "",
  "monthly_data": [
    {"sr_no": 2, "month": "February 2025", "units": 99, "bill_amount": ""},
    {"sr_no": 13, "month": "January 2026", "units": 25, "bill_amount": "320.45"}
  ],
  "average_units": "165.75",
  "average_bill_amount": "320.45",
  "kw": "1.562785714",
  "solar_panels": "2.604642857",
  "solar_capacity": "1.8",
  "number_of_panels": "3",
  "total_solar_capacity": "3.6",
  "total_number_of_solar_panels": "6"
}

Extract every monthly row with sr_no, month, units, bill_amount (empty string if blank).
Return ONLY the JSON object. No markdown, no explanation."""


def extract_bill_data(image_bytes, media_type):
    client = Groq()
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    response = client.chat.completions.create(
        model="meta-llama/llama-4-scout-17b-16e-instruct",
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{b64}"}},
                {"type": "text", "text": "Extract all data from this electricity bill."}
            ]}
        ],
        temperature=0,
        max_tokens=2000,
    )
    raw = response.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def make_border(color="000000", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value="", bold=False, fill=None,
             align="center", font_color="000000", size=11, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=font_color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border or make_border()
    if fill:
        c.fill = fill
    return c


def create_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Data"

    # ── Fills ──────────────────────────────────────────────────────────────
    orange_fill  = PatternFill("solid", fgColor="F4A023")   # header / label
    salmon_fill  = PatternFill("solid", fgColor="F4A87C")   # solar capacity / num panels label
    yellow_fill  = PatternFill("solid", fgColor="FFFF00")   # solar panel used value, solar capacity value
    green_fill   = PatternFill("solid", fgColor="92D050")   # number of panels value
    blue_fill    = PatternFill("solid", fgColor="DDEEFF")   # info label cells
    white_fill   = PatternFill("solid", fgColor="FFFFFF")

    row = 1

    # ── Info rows ──────────────────────────────────────────────────────────
    info_rows = [
        ("Consumer Name",    data.get("consumer_name", "")),
        ("Consumer No",      data.get("consumer_no", "")),
        ("Fixed Charges",    data.get("fixed_charges", "")),
        ("Sanct. Load (kW)", data.get("sanct_load_kw", "")),
        ("Connection Type",  data.get("connection_type", "")),
    ]
    for label, value in info_rows:
        ws.merge_cells(f"A{row}:B{row}")
        set_cell(ws, row, 1, label, bold=True, fill=blue_fill, align="left")
        ws.merge_cells(f"C{row}:D{row}")
        set_cell(ws, row, 3, value, align="right")
        row += 1

    # Contract Demand
    ws.merge_cells(f"A{row}:D{row}")
    set_cell(ws, row, 1, f"Contract Demand (KVA) : {data.get('contract_demand_kva','')}", align="left")
    row += 1

    # Solar Panel used
    ws.merge_cells(f"A{row}:B{row}")
    set_cell(ws, row, 1, "Solar Pannel used", bold=True, align="left")
    ws.merge_cells(f"C{row}:D{row}")
    set_cell(ws, row, 3, data.get("solar_panel_used", ""), fill=yellow_fill, align="center")
    row += 1

    # ── Table header ───────────────────────────────────────────────────────
    for col, h in enumerate(["Sr.No", "Month", "Units", "Bill Amount"], start=1):
        set_cell(ws, row, col, h, bold=True, fill=orange_fill,
                 font_color="FFFFFF", align="center")
    row += 1

    # ── Monthly rows ───────────────────────────────────────────────────────
    for entry in data.get("monthly_data", []):
        set_cell(ws, row, 1, entry.get("sr_no", ""),    align="center")
        set_cell(ws, row, 2, entry.get("month", ""),    align="right")
        set_cell(ws, row, 3, entry.get("units", ""),    align="right")
        amt = entry.get("bill_amount", "")
        set_cell(ws, row, 4, amt if amt else "",         align="right")
        row += 1

    # ── Average ────────────────────────────────────────────────────────────
    set_cell(ws, row, 1, "", align="center")
    set_cell(ws, row, 2, "Average", bold=True, align="left")
    set_cell(ws, row, 3, data.get("average_units", ""),       bold=True, align="right")
    set_cell(ws, row, 4, data.get("average_bill_amount", ""), bold=True, align="right")
    row += 1

    # ── kW ─────────────────────────────────────────────────────────────────
    set_cell(ws, row, 1, "", align="center")
    set_cell(ws, row, 2, "kW", bold=True, align="left")
    set_cell(ws, row, 3, data.get("kw", ""), align="right")
    set_cell(ws, row, 4, "", align="center")
    row += 1

    # ── Solar Panels ───────────────────────────────────────────────────────
    set_cell(ws, row, 1, "", align="center")
    set_cell(ws, row, 2, "Solar Panels", bold=True, align="left")
    set_cell(ws, row, 3, data.get("solar_panels", ""), align="right")
    set_cell(ws, row, 4, "", align="center")
    row += 1

    # ── Solar capacity (salmon label, yellow value) ────────────────────────
    set_cell(ws, row, 1, "", align="center")
    set_cell(ws, row, 2, "Solar capacity", bold=True, fill=salmon_fill,
             font_color="000000", align="left")
    set_cell(ws, row, 3, data.get("solar_capacity", ""), fill=yellow_fill, align="right")
    set_cell(ws, row, 4, "", align="center")
    row += 1

    # ── Number of Panels (salmon label, green value) ───────────────────────
    set_cell(ws, row, 1, "", align="center")
    set_cell(ws, row, 2, "Number of Panels", bold=True, fill=salmon_fill,
             font_color="000000", align="left")
    set_cell(ws, row, 3, data.get("number_of_panels", ""), fill=green_fill, align="right")
    set_cell(ws, row, 4, "", align="center")
    row += 1

    # blank row
    row += 1

    # ── Total solar capacity ───────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    set_cell(ws, row, 1, "Total solar capacity", bold=False, align="left")
    ws.merge_cells(f"C{row}:D{row}")
    set_cell(ws, row, 3, data.get("total_solar_capacity", ""), align="right")
    row += 1

    # ── Number of solar panels ─────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    set_cell(ws, row, 1, "Number of solar panels", bold=False, align="left")
    ws.merge_cells(f"C{row}:D{row}")
    set_cell(ws, row, 3, data.get("total_number_of_solar_panels", ""), align="right")

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sidebar ────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("🔑 Setup")
    api_key = st.text_input("Groq API Key", type="password", placeholder="gsk_...")
    st.caption("Get free key at [console.groq.com](https://console.groq.com)")
    if api_key:
        os.environ["GROQ_API_KEY"] = api_key
        st.success("API key set!")

# ── Main UI ────────────────────────────────────────────────────────────────
uploaded = st.file_uploader("Upload electricity bill image", type=["jpg", "jpeg", "png", "webp"])

if uploaded:
    st.image(uploaded, caption="Uploaded Bill", use_container_width=True)

    if st.button("🔍 Extract Data", type="primary"):
        if not api_key:
            st.warning("⚠️ Please enter your Groq API key in the sidebar first.")
        else:
            with st.spinner("AI is reading your bill..."):
                try:
                    image_bytes = uploaded.read()
                    data = extract_bill_data(image_bytes, uploaded.type)
                    st.success("✅ Data extracted successfully!")
                    st.session_state["extracted_data"] = data
                    st.session_state["filename"] = uploaded.name
                except Exception as e:
                    st.error(f"❌ Error: {e}")

if "extracted_data" in st.session_state:
    data = st.session_state["extracted_data"]

    # Info preview
    st.subheader("📋 Consumer Info")
    c1, c2 = st.columns(2)
    with c1:
        st.write(f"**Consumer Name:** {data.get('consumer_name','')}")
        st.write(f"**Consumer No:** {data.get('consumer_no','')}")
        st.write(f"**Fixed Charges:** {data.get('fixed_charges','')}")
    with c2:
        st.write(f"**Sanct. Load:** {data.get('sanct_load_kw','')}")
        st.write(f"**Connection Type:** {data.get('connection_type','')}")
        st.write(f"**Solar Panel Used:** {data.get('solar_panel_used','')}")

    # Monthly table
    st.subheader("📊 Monthly Consumption")
    monthly = data.get("monthly_data", [])
    if monthly:
        st.dataframe(
            {"Sr.No":       [m.get("sr_no","")       for m in monthly],
             "Month":       [m.get("month","")        for m in monthly],
             "Units":       [m.get("units","")        for m in monthly],
             "Bill Amount": [m.get("bill_amount","")  for m in monthly]},
            use_container_width=True, hide_index=True
        )

    # Solar summary
    st.subheader("☀️ Solar Summary")
    c1, c2 = st.columns(2)
    with c1:
        st.write(f"**Solar Capacity:** {data.get('solar_capacity','')}")
        st.write(f"**Number of Panels:** {data.get('number_of_panels','')}")
    with c2:
        st.write(f"**Total Solar Capacity:** {data.get('total_solar_capacity','')}")
        st.write(f"**Total Solar Panels:** {data.get('total_number_of_solar_panels','')}")

    # Download
    excel_buf = create_excel(data)
    filename = st.session_state.get("filename", "bill").rsplit(".", 1)[0]
    st.download_button(
        label="⬇️ Download as Excel",
        data=excel_buf,
        file_name=f"{filename}_extracted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )